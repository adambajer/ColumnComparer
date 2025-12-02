import os
import csv
import warnings
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# potlačení warningů z openpyxl
warnings.filterwarnings("ignore")

from openpyxl import load_workbook


# =====================================================================
# Pomocné funkce – načítání CSV / Excel BEZ pandas
# =====================================================================

def read_csv_table(path):
    """
    Načte CSV/TXT a vrátí:
    - list názvů sloupců
    - list řádků (dict column → value)
    """
    # autodetekce oddělovače na vzorku
    with open(path, "r", encoding="utf-8-sig") as f:
        sample = f.read(4096)

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample)
    except csv.Error:
        dialect = csv.excel

    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, dialect=dialect)
        headers = reader.fieldnames or []
        for row in reader:
            rows.append({k: (v if v is not None else "") for k, v in row.items()})

    return headers, rows


def read_excel_table(path):
    """
    Načte .xlsx BEZ pandas.
    Vrací (headers, rows) – headers = list názvů sloupců, rows = list dictů.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if not wb.sheetnames:
        wb.close()
        raise RuntimeError("Excel neobsahuje žádné listy.")

    ws = wb[wb.sheetnames[0]]

    rows = []
    headers = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(x) if x is not None else "" for x in row]
            continue
        rows.append({
            headers[j]: (str(v) if v is not None else "")
            for j, v in enumerate(row)
        })

    wb.close()
    return headers, rows


def read_table(path):
    """
    Detekuje typ souboru a načte ho jako tabulku.
    Podporuje: .xlsx, .xlsm, .csv, .txt
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return read_excel_table(path)

    if ext in (".csv", ".txt"):
        return read_csv_table(path)

    raise RuntimeError(f"Nepodporovaný formát souboru: {ext}")


# =====================================================================
# GUI aplikace
# =====================================================================

class ColumnComparerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Porovnávač sloupců — ULTRA LITE verze")
        self.geometry("1400x800")

        # data tabulky 1
        self.headers1 = []
        self.rows1 = []

        # data tabulky 2
        self.headers2 = []
        self.rows2 = []

        # vybraný sloupec z tabulky 1 a 2
        self.col1 = None
        self.col2 = None

        # aktuální výsledek (pro export)
        self.result_rows = []

        self._build_ui()

    # ------------------------------------------------------------------
    def _build_ui(self):
        paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, pady=5)

        # ======================== TABULKA 1 ============================
        left = ttk.Labelframe(paned, text="Tabulka 1")
        paned.add(left, weight=1)

        lf = ttk.Frame(left)
        lf.pack(fill=tk.X)
        ttk.Button(lf, text="Načíst tabulku 1", command=self.load_table1).pack(
            side=tk.LEFT, padx=5
        )
        self.lbl_file1 = ttk.Label(lf, text="(není načteno)")
        self.lbl_file1.pack(side=tk.LEFT, padx=10)

        self.tree1 = self._create_tree(left, self.on_col1)

        # ======================== TABULKA 2 ============================
        right = ttk.Labelframe(paned, text="Tabulka 2")
        paned.add(right, weight=1)

        rf = ttk.Frame(right)
        rf.pack(fill=tk.X)
        ttk.Button(rf, text="Načíst tabulku 2", command=self.load_table2).pack(
            side=tk.LEFT, padx=5
        )
        self.lbl_file2 = ttk.Label(rf, text="(není načteno)")
        self.lbl_file2.pack(side=tk.LEFT, padx=10)

        self.tree2 = self._create_tree(right, self.on_col2)

        # ====================== Spodní ovládací panel ==================
        bottom = ttk.Frame(self)
        bottom.pack(fill=tk.X, pady=8)

        self.lbl_sel1 = ttk.Label(bottom, text="Sloupec 1: (nic)")
        self.lbl_sel1.pack(side=tk.LEFT, padx=10)

        self.lbl_sel2 = ttk.Label(bottom, text="Sloupec 2: (nic)")
        self.lbl_sel2.pack(side=tk.LEFT, padx=10)

        ttk.Button(
            bottom,
            text="Porovnat – shody",
            command=self.compare_matches
        ).pack(side=tk.LEFT, padx=10)

        ttk.Button(
            bottom,
            text="Porovnat – rozdíly",
            command=self.compare_differences
        ).pack(side=tk.LEFT, padx=10)

        ttk.Button(
            bottom,
            text="Export výsledku",
            command=self.export
        ).pack(side=tk.LEFT, padx=10)

        # ========================= Výsledek ============================
        result_frame = ttk.Labelframe(self, text="Výsledné řádky (Tabulka 1)")
        result_frame.pack(fill=tk.BOTH, expand=True)

        self.tree_result = self._create_tree(result_frame, None)

    # ------------------------------------------------------------------
    def _create_tree(self, parent, header_callback):
        """
        Vytvoří TreeView s posuvníky a volitelným callbackem na kliknutí hlavičky.
        """
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.BOTH, expand=True)

        tree = ttk.Treeview(frame, show="headings")
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        if header_callback:
            def handler(event, tv=tree, cb=header_callback):
                region = tv.identify_region(event.x, event.y)
                if region != "heading":
                    return
                col_id = tv.identify_column(event.x)  # "#1", "#2", ...
                try:
                    idx = int(col_id[1:]) - 1
                except ValueError:
                    return
                cols = list(tv["columns"])
                if 0 <= idx < len(cols):
                    cb(cols[idx])
            tree.bind("<Button-1>", handler, add="+")

        return tree

    # ------------------------------------------------------------------
    # Načítání tabulek
    # ------------------------------------------------------------------
    def load_table1(self):
        path = filedialog.askopenfilename(
            title="Vyber tabulku 1",
            filetypes=[("Tabulky", "*.xlsx *.xlsm *.csv *.txt"), ("Vše", "*.*")]
        )
        if not path:
            return

        try:
            self.headers1, self.rows1 = read_table(path)
            self.lbl_file1.config(text=os.path.basename(path))
            self.fill_tree(self.tree1, self.headers1, self.rows1)
        except Exception as e:
            messagebox.showerror("Chyba při načítání tabulky 1", str(e))

    def load_table2(self):
        path = filedialog.askopenfilename(
            title="Vyber tabulku 2",
            filetypes=[("Tabulky", "*.xlsx *.xlsm *.csv *.txt"), ("Vše", "*.*")]
        )
        if not path:
            return

        try:
            self.headers2, self.rows2 = read_table(path)
            self.lbl_file2.config(text=os.path.basename(path))
            self.fill_tree(self.tree2, self.headers2, self.rows2)
        except Exception as e:
            messagebox.showerror("Chyba při načítání tabulky 2", str(e))

    # ------------------------------------------------------------------
    def fill_tree(self, tree, headers, rows):
        """
        Naplní TreeView daty.
        """
        tree.delete(*tree.get_children())
        tree["columns"] = headers

        for h in headers:
            tree.heading(h, text=h)
            tree.column(h, width=120, stretch=True)

        for row in rows:
            tree.insert("", tk.END, values=[row.get(h, "") for h in headers])

    # ------------------------------------------------------------------
    # Výběr sloupců kliknutím na hlavičku
    # ------------------------------------------------------------------
    def on_col1(self, col):
        self.col1 = col
        self.lbl_sel1.config(text=f"Sloupec 1: {col}")

    def on_col2(self, col):
        self.col2 = col
        self.lbl_sel2.config(text=f"Sloupec 2: {col}")

    # ------------------------------------------------------------------
    # Porovnání – SHODY
    # ------------------------------------------------------------------
    def compare_matches(self):
        """
        Najde řádky z tabulky 1, jejichž hodnota ve sloupci col1
        se VYSKYTUJE ve sloupci col2 tabulky 2.
        """
        if not self.col1 or not self.col2:
            messagebox.showwarning(
                "Chyba",
                "Vyber oba sloupce kliknutím na hlavičky (Tabulka 1 i Tabulka 2)."
            )
            return

        values2 = {row[self.col2] for row in self.rows2}

        self.result_rows = [
            row for row in self.rows1
            if row[self.col1] in values2
        ]

        self.fill_tree(self.tree_result, self.headers1, self.result_rows)

        messagebox.showinfo(
            "Hotovo",
            f"Nalezeno {len(self.result_rows)} řádků se SHODNOU hodnotou."
        )

    # ------------------------------------------------------------------
    # Porovnání – ROZDÍLY (opak)
    # ------------------------------------------------------------------
    def compare_differences(self):
        """
        Najde řádky z tabulky 1, jejichž hodnota ve sloupci col1
        se NEVYSKYTUJE ve sloupci col2 tabulky 2.
        """
        if not self.col1 or not self.col2:
            messagebox.showwarning(
                "Chyba",
                "Vyber oba sloupce kliknutím na hlavičky (Tabulka 1 i Tabulka 2)."
            )
            return

        values2 = {row[self.col2] for row in self.rows2}

        self.result_rows = [
            row for row in self.rows1
            if row[self.col1] not in values2
        ]

        self.fill_tree(self.tree_result, self.headers1, self.result_rows)

        messagebox.showinfo(
            "Hotovo",
            f"Nalezeno {len(self.result_rows)} řádků, které NEMAJÍ hodnotu z tabulky 2."
        )

    # ------------------------------------------------------------------
    # Export výsledku
    # ------------------------------------------------------------------
    def export(self):
        if not self.result_rows:
            messagebox.showwarning("Nic k exportu", "Nejdřív proveď porovnání.")
            return

        path = filedialog.asksaveasfilename(
            title="Uložit výsledek",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Text", "*.txt")]
        )
        if not path:
            return

        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=self.headers1)
                writer.writeheader()
                writer.writerows(self.result_rows)

            messagebox.showinfo("Export", f"Výsledek uložen do:\n{path}")
        except Exception as e:
            messagebox.showerror("Chyba při exportu", str(e))


# =====================================================================
if __name__ == "__main__":
    ColumnComparerApp().mainloop()
